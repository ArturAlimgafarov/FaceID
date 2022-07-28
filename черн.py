import cv2
import dlib
import openpyxl as xlsx
import math
import asyncio

async def getName(image, database):
    sp = dlib.shape_predictor('shape_predictor_68_face_landmarks.dat')  # 'shape_predictor_5_face_landmarks.dat'
    face_rec = dlib.face_recognition_model_v1('dlib_face_recognition_resnet_model_v1.dat')
    detector = dlib.get_frontal_face_detector()

    bio = None
    try:
        detected_face = detector(image, 1)[0]
        shape = sp(image, detected_face)
        face_descriptor = face_rec.compute_face_descriptor(image, shape)
        if (face_descriptor != None):
            bio = face_descriptor
    except:
        bio = None
    username = 'unknown'
    if bio != None:
        for data in database:
            if len(data) == 129:
                if math.dist(bio, data[1:]) < 0.6:
                    username = data[0]
                    break
    print(username)
def getDB(dbFilename):
    wb = xlsx.load_workbook(dbFilename)
    sheet = wb.active
    dataCount = 128

    db = []
    rowIndex = 1
    while sheet.cell(rowIndex, 1).value:
        data = [sheet.cell(rowIndex, 1).value]
        for j in range(2, dataCount + 2):
            data.append(sheet.cell(rowIndex, j).value)
        rowIndex += 1
        db.append(data)
    return db

db = getDB('data.xlsx')
capture = cv2.VideoCapture(0, cv2.CAP_DSHOW)
face_cscd = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

while True:
    _, img = capture.read()
    faces = face_cscd.detectMultiScale(img, scaleFactor=1.5, minNeighbors=5, minSize=(50, 50))

    loop = asyncio.get_event_loop()
    loop.run_until_complete(
        asyncio.wait([getName(img, db)])
    )
    for face in faces:
        (x, y, w, h) = face
        cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 255), 2)

    cv2.imshow('Camera', img)

    key = cv2.waitKey(2) & 0xFF
    if key == 27:
        break

capture.release()
cv2.destroyAllWindows()